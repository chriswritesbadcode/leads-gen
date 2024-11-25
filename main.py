import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
import requests
import json

extractionsFolder = './Extractions'
outputFile = 'Leads.xlsx'
apiUrl = 'https://cloud.gmapsextractor.com/api/v1/search'
apiToken = ''
keepColumns = ['Name', 'Municipality', 'Categories', 'Time Zone', 'Phone', 'Claimed', 'Review Count', 'Average Rating', 'Website'] 

allData = pd.DataFrame()

for file in os.listdir(extractionsFolder):
    filePath = os.path.join(extractionsFolder, file)
    df = pd.read_csv(filePath)
    dfFiltered = df[keepColumns]
    allData = pd.concat([allData, dfFiltered], ignore_index=True)

if os.path.exists(outputFile):
    existingData = pd.read_excel(outputFile)
    allData = pd.concat([existingData, allData], ignore_index=True)

allData = allData.drop_duplicates(subset=['Name', 'Phone'], keep='first')

allData = allData.sort_values(by='Time Zone')

if 'Called' not in allData.columns:
    allData.insert(0, 'Called', '')

allData.rename(columns={'Review Count': 'Reviews', 'Average Rating': 'Rating'}, inplace=True)

allData.to_excel(outputFile, index=False)

wb = load_workbook(outputFile)
ws = wb.active

headerFill = PatternFill(start_color='ADDFFF', end_color='ADDFFF', fill_type='solid')
for cell in ws[1]:
    cell.fill = headerFill
    cell.alignment = Alignment(horizontal='center', vertical='center')

columnStyles = {
    "A": {"width": 10, "alignment": Alignment(horizontal="center", vertical="center")}, # Contacted
    "B": {"width": 50, "alignment": Alignment(horizontal="center", vertical="center")}, # Name
    "C": {"width": 25, "alignment": Alignment(horizontal="center", vertical="center")}, # Municipality
    "D": {"width": 25, "alignment": Alignment(horizontal="center", vertical="center")}, # Category
    "E": {"width": 20, "alignment": Alignment(horizontal="center", vertical="center")}, # Time Zone
    "F": {"width": 25, "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)}, # Phone
    "G": {"width": 10, "alignment": Alignment(horizontal="center", vertical="center")},  # Claimed
    "H": {"width": 10, "alignment": Alignment(horizontal="center", vertical="center")},  # Reviews
    "I": {"width": 10, "alignment": Alignment(horizontal="center", vertical="center")},  # Rating
    "J": {"width": 50, "alignment": Alignment(horizontal="left", vertical="center")},  # Website
}

for columnLetter, style in columnStyles.items():
    for cell in ws[columnLetter]:
        cell.alignment = style['alignment']
    ws.column_dimensions[columnLetter].width = style['width']

for cell in ws['F']:
    if cell.row > 1 and cell.value:
        cell.value = cell.value.split(',')[1] if ',' in cell.value else cell.value

wb.save(outputFile)

print(f'Leads successfully merged into {outputFile}.')